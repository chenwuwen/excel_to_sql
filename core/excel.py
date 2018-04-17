import openpyxl
import os

base_dir = 'E:\\'

titles = ['地址区县', '街路巷名称', '门楼牌号地址', '单元名称', '户名称', '层数', '坐标x', '坐标y']
src_filename = "dzys.xlsx"
filename = "data.txt"
script = "script.txt"

# 当前文件的决对路径,包括文件名
current_abs_path = os.path.abspath(__file__)
current_path = os.path.dirname(__file__)
print(current_path)


# 获取excel内容
def get_excel_data():
    excel = openpyxl.load_workbook(base_dir + src_filename)
    print(excel.sheetnames)
    sheet = excel['Sheet1']

    # for i,row in enumerate(sheet.values):
    #     print(i,row)

    for row in sheet.iter_rows(min_row=3):
        # print(row)
        # print()
        for i, cell in enumerate(row):
            pass
            # print('当前列的标题：',sheet.cell(row=1, column=i + 1).value)
            # columnName = cell.column
            # print('当前列的名称为：', columnName)
            # print(cell.value, end='\t')
    return sheet


# 将excel内容导出到文本
def excel_to_txt(sheet):
    with open(filename, 'w+', encoding='utf-8') as data:
        for row in sheet.iter_rows(min_row=3):
            txt = []
            # list = []
            for i, cell in enumerate(row):
                title = sheet.cell(row=1, column=i + 1).value.strip()
                if title in titles:
                    # list.append(title)
                    txt.append(cell.value)
            content = "\t".join(txt)
            content += "\n"
            print(content)
            data.writelines(content)


# 得到已排列好的嵌套列表
def read_file_tolist():
    with open(current_path + "/" + filename, 'r', encoding='utf-8') as data:
        print("############################################################")
        big_list = []
        for line in data.readlines():
            print(line)
            smal_list = line.split("\t")
            sort_list = list(range(9))
            # 户名称
            sort_list[6] = smal_list[4].strip()
            # 层数
            sort_list[5] = smal_list[5].strip()
            # 单元名称
            sort_list[4] = smal_list[3].strip()
            # 楼名称
            sort_list[3] = smal_list[2].strip()
            sort_list[0] = smal_list[0].strip()
            # 经度
            sort_list[7] = smal_list[6].strip()
            # 纬度
            sort_list[8] = smal_list[7].strip()
            # 分割 镇/街道 路 小区
            z_l_x = smal_list[1].strip()
            x, y = split_str(z_l_x)
            sort_list[1] = x.strip()
            sort_list[2] = y.strip()
            big_list.append(sort_list)
    return big_list


def list_to_file(big_list):
    with open(script, 'w+', encoding='utf-8') as data:
        content = ""
        for sort_list in big_list:
            content = "\t".join(sort_list)
            content += "\n"
            data.writelines(content)


# 分割字符串
def split_str(str):
    # 检测 str是否包含在字符串中，如果指定范围beg和end ，则检查是否包含在指定范围内，如果包含返回开始的索引值，否则返回 - 1
    s = str.find("镇")
    k = str.find("街道")
    x = ""
    y = ""
    # 如果这个字符串内同时存在"镇","街道",以"街道"为准
    k, s = int(k), int(s)
    if k > 0 and s > 0:
        x = str[0:k + 1]
        y = str[k + 1:]
    else:
        if k > 0:
            x = str[0:k + 1]
            y = str[k + 1:]
        else:
            x = str[0:s + 1]
            y = str[s + 1:]
    # 返回x为"路",y为小区
    return x, y


# 执行顺序
# sheet = get_excel_data()
# excel_to_txt(sheet)
# big_list = read_file_tolist()
# list_to_file(big_list)


def exec():
    sheet = get_excel_data()
    excel_to_txt(sheet)
    big_list = read_file_tolist()
    list_to_file(big_list)


# 这个主要是用来防止其他模块调用本模块的方法时，执行两次(先执行自己的，在执行调用的)的方法，只有单独运行该文件时才会走下面的方法
if __name__ == '__main__':
    exec()
